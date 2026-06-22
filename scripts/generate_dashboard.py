#!/usr/bin/env python3
from __future__ import annotations

import datetime as dt
import html
import json
import math
import re
import statistics
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
TEMPLATE = ROOT / "templates" / "dashboard-template.html"
OUTPUT = ROOT / "public" / "index.html"

ROSTER_FILE = DATA_DIR / "HR台账模板_员工花名册.xlsx"
SALARY_FILE = DATA_DIR / "HR台账模板_薪资数据.xlsx"
PERF_FILE = DATA_DIR / "HR台账模板_绩效数据.xlsx"


def today() -> dt.date:
    return dt.date.today()


def text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def number(value, default=0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        if math.isnan(value):
            return default
        return float(value)
    raw = str(value).replace(",", "").replace("¥", "").replace("￥", "").strip()
    if not raw:
        return default
    try:
        return float(raw)
    except ValueError:
        return default


def parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    raw = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return dt.datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    return None


def pct(numerator, denominator) -> float:
    if not denominator:
        return 0.0
    return numerator / denominator * 100


def fmt_pct(value, digits=1) -> str:
    return f"{value:.{digits}f}%"


def fmt_money(value) -> str:
    return f"¥{value:,.0f}"


def fmt_num(value, digits=1) -> str:
    if value is None:
        return "0"
    if abs(value - round(value)) < 1e-9:
        return str(int(round(value)))
    return f"{value:.{digits}f}"


def table(headers, rows) -> str:
    head = "".join(f"<th>{html.escape(str(h))}</th>" for h in headers)
    body = []
    for row in rows:
        body.append("<tr>" + "".join(f"<td>{html.escape(str(c))}</td>" for c in row) + "</tr>")
    return '<table class="dt"><tr>' + head + "</tr>" + "".join(body) + "</table>"


def load_rows(path: Path) -> list[dict[str, object]]:
    if not path.exists():
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    header = None
    data = []
    for row in sheet.iter_rows(values_only=True):
        values = list(row)
        if header is None:
            if any(text(v) == "员工ID" for v in values):
                header = [text(v) for v in values]
            continue
        if not any(text(v) for v in values):
            continue
        first = text(values[0])
        if first.startswith("...") or first.startswith("（") or first.startswith("("):
            continue
        item = {header[i]: values[i] if i < len(values) else None for i in range(len(header))}
        if text(item.get("员工ID")):
            data.append(item)
    return data


def normalize_gender(value: str) -> str:
    raw = value.lower()
    if raw in {"male", "m", "男"}:
        return "男"
    if raw in {"female", "f", "女"}:
        return "女"
    return value or "未知"


def normalize_edu(value: str) -> str:
    mapping = {
        "phd": "博士",
        "doctor": "博士",
        "doctoral": "博士",
        "master": "硕士",
        "bachelor": "本科",
        "college": "大专",
        "high school": "高中及以下",
    }
    raw = value.strip()
    return mapping.get(raw.lower(), raw or "其他")


def normalize_city(value: str) -> str:
    mapping = {
        "shenzhen": "深圳",
        "beijing": "北京",
        "shanghai": "上海",
        "guangzhou": "广州",
    }
    return mapping.get(value.lower(), value or "未知")


def normalize_dept(value: str) -> str:
    mapping = {
        "r&d center": "研发中心",
        "product mgmt": "产品管理部",
        "testing dept": "测试部",
        "admin dept": "综合管理部",
        "operations support": "运维支持部",
    }
    return mapping.get(value.lower(), value or "未填部门")


def normalize_position(value: str) -> str:
    mapping = {
        "senior engineer": "高级工程师",
        "product manager": "产品经理",
        "architect": "架构师",
        "hr bp": "HR BP",
        "test engineer": "测试工程师",
    }
    return mapping.get(value.lower(), value or "未填岗位")


def normalize_role(value: str) -> str:
    mapping = {
        "tech expert": "技术专家",
        "key role": "关键岗位",
        "regular": "普通员工",
    }
    return mapping.get(value.lower(), value or "普通员工")


def normalize_marriage(value: str) -> str:
    mapping = {
        "single": "未婚",
        "married w/ kids": "已婚已育",
        "married with kids": "已婚已育",
        "married no kids": "已婚未育",
        "married without kids": "已婚未育",
    }
    return mapping.get(value.lower(), value or "其他")


def normalize_status(value: str) -> str:
    raw = value.lower().replace(" ", "")
    if raw in {"resigned", "离职", "已离职"}:
        return "resigned"
    if raw in {"resigning", "离职中", "待离职"}:
        return "active"
    return "active"


def is_key_role(employee: dict) -> bool:
    blob = (employee.get("role", "") + " " + employee.get("position", "")).lower()
    keys = ["关键", "核心", "骨干", "key", "core", "manager", "主管", "项目经理", "负责人", "lead"]
    return any(k in blob for k in keys)


def band_age(age: float) -> str:
    if age < 25:
        return "20-25岁"
    if age <= 30:
        return "26-30岁"
    if age <= 35:
        return "31-35岁"
    if age <= 40:
        return "36-40岁"
    return "40岁+"


def band_tenure(years: float) -> str:
    if years < 1:
        return "0-1年"
    if years <= 3:
        return "1-3年"
    if years <= 5:
        return "3-5年"
    if years <= 10:
        return "5-10年"
    return "10年+"


def band_salary(value: float) -> str:
    if value < 15000:
        return "<1.5万"
    if value < 25000:
        return "1.5-2.5万"
    if value < 35000:
        return "2.5-3.5万"
    if value < 50000:
        return "3.5-5万"
    return "5万+"


def counts(labels, counter: Counter) -> dict:
    return {"l": list(labels), "d": [counter.get(label, 0) for label in labels]}


def top_counter(counter: Counter, limit=8) -> dict:
    items = counter.most_common(limit)
    if not items:
        return {"l": ["暂无数据"], "d": [0]}
    return {"l": [k for k, _ in items], "d": [v for _, v in items]}


def top_values(mapping: dict[str, float], limit=8) -> dict:
    items = sorted(mapping.items(), key=lambda kv: kv[1], reverse=True)[:limit]
    if not items:
        return {"l": ["暂无数据"], "d": [0]}
    return {"l": [k for k, _ in items], "d": [round(v, 1) for _, v in items]}


def avg(values: list[float]) -> float:
    values = [v for v in values if v is not None]
    return sum(values) / len(values) if values else 0.0


def median(values: list[float]) -> float:
    values = sorted(v for v in values if v is not None)
    return statistics.median(values) if values else 0.0


def percentile(values: list[float], p: float) -> float:
    values = sorted(v for v in values if v is not None)
    if not values:
        return 0.0
    if len(values) == 1:
        return values[0]
    k = (len(values) - 1) * p
    lo = math.floor(k)
    hi = math.ceil(k)
    if lo == hi:
        return values[int(k)]
    return values[lo] * (hi - k) + values[hi] * (k - lo)


def group_average(rows, key_field, value_field):
    groups = defaultdict(list)
    for row in rows:
        key = row.get(key_field) or "未分类"
        groups[key].append(row.get(value_field, 0))
    return {k: avg(v) for k, v in groups.items()}


def normalize_roster(rows):
    out = []
    report_date = today()
    for row in rows:
        hire_date = parse_date(row.get("入职日期"))
        leave_date = parse_date(row.get("离职日期") or row.get("离职时间") or row.get("离职日"))
        birth_date = parse_date(row.get("出生日期"))
        contract_end = parse_date(row.get("合同到期日"))
        confirm_date = parse_date(row.get("转正日期"))
        status = normalize_status(text(row.get("在职状态")))
        if leave_date and leave_date <= report_date:
            status = "resigned"
        age = ((report_date - birth_date).days / 365.25) if birth_date else 0
        tenure = ((report_date - hire_date).days / 365.25) if hire_date else 0
        employee = {
            "id": text(row.get("员工ID")),
            "name": text(row.get("姓名")),
            "gender": normalize_gender(text(row.get("性别"))),
            "birth_date": birth_date,
            "age": age,
            "edu": normalize_edu(text(row.get("学历"))),
            "marriage": normalize_marriage(text(row.get("婚育状态"))),
            "school": text(row.get("毕业院校")) or "未填院校",
            "dept": normalize_dept(text(row.get("部门"))),
            "position": normalize_position(text(row.get("岗位"))),
            "role": normalize_role(text(row.get("岗位角色"))),
            "city": normalize_city(text(row.get("工作城市"))),
            "hire_date": hire_date,
            "leave_date": leave_date,
            "confirm_date": confirm_date,
            "contract_end": contract_end,
            "status": status,
        }
        employee["key_role"] = is_key_role(employee)
        employee["tenure"] = max(0, tenure)
        out.append(employee)
    return out


def normalize_salary(rows):
    out = []
    for row in rows:
        wage = number(row.get("合同工资(月)"))
        bonus = number(row.get("月奖金"))
        out.append(
            {
                "id": text(row.get("员工ID")),
                "name": text(row.get("姓名")),
                "dept": normalize_dept(text(row.get("部门"))),
                "position": normalize_position(text(row.get("岗位"))),
                "edu": normalize_edu(text(row.get("学历"))),
                "city": normalize_city(text(row.get("工作城市"))),
                "gender": normalize_gender(text(row.get("性别"))),
                "wage": wage,
                "bonus": bonus,
                "total": wage + bonus,
                "range_low": number(row.get("薪酬带宽下限")),
                "range_high": number(row.get("薪酬带宽上限")),
            }
        )
    return [r for r in out if r["id"]]


def normalize_perf(rows):
    latest_by_employee = {}
    for row in rows:
        employee_id = text(row.get("员工ID"))
        if not employee_id:
            continue
        cycle = text(row.get("绩效周期"))
        item = {
            "id": employee_id,
            "name": text(row.get("姓名")),
            "dept": normalize_dept(text(row.get("部门"))),
            "position": normalize_position(text(row.get("岗位"))),
            "grade": text(row.get("绩效等级")) or "未评级",
            "score": number(row.get("绩效量化分"), None),
            "cycle": cycle,
        }
        old = latest_by_employee.get(employee_id)
        if old is None or cycle >= old.get("cycle", ""):
            latest_by_employee[employee_id] = item
    return list(latest_by_employee.values())


def build_dashboard():
    report_date = today()
    report_month = f"{report_date.year}年{report_date.month}月"
    roster = normalize_roster(load_rows(ROSTER_FILE))
    active = [e for e in roster if e["status"] == "active"]
    resigned = [e for e in roster if e["status"] == "resigned"]
    active_ids = {e["id"] for e in active}

    salaries = normalize_salary(load_rows(SALARY_FILE))
    salaries = [s for s in salaries if s["id"] in active_ids]
    salary_by_id = {s["id"]: s for s in salaries}

    perfs = normalize_perf(load_rows(PERF_FILE))
    perfs = [p for p in perfs if p["id"] in active_ids]
    perf_by_id = {p["id"]: p for p in perfs}

    active_count = len(active)
    join_this_month = [
        e for e in active if e["hire_date"] and e["hire_date"].year == report_date.year and e["hire_date"].month == report_date.month
    ]
    leave_this_month = [
        e for e in resigned if e.get("leave_date") and e["leave_date"].year == report_date.year and e["leave_date"].month == report_date.month
    ]
    join_count = len(join_this_month)
    leave_count = len(leave_this_month)
    net_change = join_count - leave_count
    avg_headcount = active_count + leave_count / 2
    turnover_rate = pct(leave_count, avg_headcount)

    ages = [e["age"] for e in active if e["age"]]
    tenures = [e["tenure"] for e in active if e["tenure"] is not None]
    avg_age = avg(ages)
    avg_tenure = avg(tenures)
    med_tenure = median(tenures)

    gender_counter = Counter(e["gender"] for e in active)
    edu_counter = Counter(e["edu"] for e in active)
    age_counter = Counter(band_age(e["age"]) for e in active if e["age"])
    tenure_counter = Counter(band_tenure(e["tenure"]) for e in active)
    dept_counter = Counter(e["dept"] for e in active)
    role_counter = Counter(e["position"] for e in active)
    role_type_counter = Counter(e["role"] for e in active)
    marriage_counter = Counter(e["marriage"] for e in active)
    school_counter = Counter(e["school"] for e in active)
    city_counter = Counter(e["city"] for e in active)

    probation = [e for e in active if not e["confirm_date"]]
    confirmed_due = [e for e in active if e["hire_date"] and (report_date - e["hire_date"]).days >= 180]
    confirmed = [e for e in confirmed_due if e["confirm_date"]]
    confirm_rate = pct(len(confirmed), len(confirmed_due))

    expired_contract = [e for e in active if e["contract_end"] and e["contract_end"] < report_date]
    contract_30 = [e for e in active if e["contract_end"] and 0 <= (e["contract_end"] - report_date).days <= 30]
    contract_90 = [e for e in active if e["contract_end"] and 30 < (e["contract_end"] - report_date).days <= 90]
    contract_warning_count = len(expired_contract) + len(contract_30) + len(contract_90)
    contract_warning_rate = pct(contract_warning_count, active_count)

    salary_values = [s["total"] for s in salaries if s["total"] > 0]
    salary_mean = avg(salary_values)
    salary_median = median(salary_values)
    salary_min = min(salary_values) if salary_values else 0
    salary_max = max(salary_values) if salary_values else 0
    salary_p25 = percentile(salary_values, 0.25)
    salary_p75 = percentile(salary_values, 0.75)
    salary_p10 = percentile(salary_values, 0.10)
    salary_p90 = percentile(salary_values, 0.90)
    salary_range_ratio = (salary_max / salary_min) if salary_min else 0
    bonuses = [s["bonus"] for s in salaries if s["bonus"] is not None]
    bonus_positive = [s for s in salaries if s["bonus"] > 0]
    bonus_mean = avg(bonuses)
    bonus_total = sum(bonuses)
    bonus_coverage_rate = pct(len(bonus_positive), active_count)
    bonus_salary_ratio = pct(bonus_mean, salary_mean)

    salary_hist = Counter(band_salary(v) for v in salary_values)
    dept_salary = group_average(salaries, "dept", "total")
    role_salary = group_average(salaries, "position", "total")
    edu_salary = group_average(salaries, "edu", "total")
    city_salary = group_average(salaries, "city", "total")
    gender_salary = group_average(salaries, "gender", "total")

    salary_by_tenure_groups = defaultdict(list)
    for employee in active:
        sal = salary_by_id.get(employee["id"])
        if sal:
            salary_by_tenure_groups[band_tenure(employee["tenure"])].append(sal["total"])
    salary_by_tenure = {k: avg(v) for k, v in salary_by_tenure_groups.items()}

    perf_counter = Counter(p["grade"] for p in perfs)
    perf_count = len(perfs)
    high_perf_count = sum(1 for p in perfs if p["grade"] in {"A", "B+"})
    high_perf_rate = pct(high_perf_count, perf_count)
    perf_coverage_rate = pct(perf_count, active_count)
    perf_scores = [p["score"] for p in perfs if p["score"] is not None]
    perf_score_mean = avg(perf_scores)

    perf_salary_groups = defaultdict(list)
    dept_perf_grade = {dept: Counter() for dept in dept_counter.keys()}
    dept_high_perf = defaultdict(lambda: [0, 0])
    role_high_perf = defaultdict(lambda: [0, 0])
    tenure_high_perf = defaultdict(lambda: [0, 0])
    for p in perfs:
        sal = salary_by_id.get(p["id"])
        emp = next((e for e in active if e["id"] == p["id"]), None)
        if sal:
            perf_salary_groups[p["grade"]].append(sal["total"])
        if emp:
            dept_perf_grade[emp["dept"]][p["grade"]] += 1
            dept_high_perf[emp["dept"]][1] += 1
            role_high_perf[emp["position"]][1] += 1
            tenure_high_perf[band_tenure(emp["tenure"])][1] += 1
            if p["grade"] in {"A", "B+"}:
                dept_high_perf[emp["dept"]][0] += 1
                role_high_perf[emp["position"]][0] += 1
                tenure_high_perf[band_tenure(emp["tenure"])][0] += 1

    one_year_people = [e for e in roster if e["hire_date"] and (report_date - e["hire_date"]).days <= 365]
    one_year_resigned = [e for e in one_year_people if e["status"] == "resigned"]
    new_hire_turnover_rate = pct(len(one_year_resigned), len(one_year_people))

    def turnover_by(field):
        group_total = Counter(e[field] for e in roster)
        group_leave = Counter(e[field] for e in resigned)
        return {k: pct(group_leave[k], v) for k, v in group_total.items()}

    dept_turnover = turnover_by("dept")
    city_turnover = turnover_by("city")
    role_turnover = turnover_by("position")
    edu_turnover = turnover_by("edu")
    tenure_turnover_total = Counter(band_tenure(e["tenure"]) for e in roster)
    tenure_turnover_leave = Counter(band_tenure(e["tenure"]) for e in resigned)
    tenure_turnover = {k: pct(tenure_turnover_leave[k], v) for k, v in tenure_turnover_total.items()}
    key_roles = [e for e in roster if e["key_role"]]
    key_role_resigned = [e for e in key_roles if e["status"] == "resigned"]
    key_role_turnover_rate = pct(len(key_role_resigned), len(key_roles))

    contract_labels = ["已过期", "30天内", "31-90天", "正常"]
    contract_data = [
        len(expired_contract),
        len(contract_30),
        len(contract_90),
        max(active_count - contract_warning_count, 0),
    ]

    age_labels = ["20-25岁", "26-30岁", "31-35岁", "36-40岁", "40岁+"]
    tenure_labels = ["0-1年", "1-3年", "3-5年", "5-10年", "10年+"]
    salary_labels = ["<1.5万", "1.5-2.5万", "2.5-3.5万", "3.5-5万", "5万+"]
    perf_labels = ["A", "B+", "B", "C"]
    dept_order = [k for k, _ in dept_counter.most_common()] or ["暂无数据"]

    cd = {
        "gender": top_counter(gender_counter),
        "tenure": counts(tenure_labels, tenure_counter),
        "age": counts(age_labels, age_counter),
        "edu": top_counter(edu_counter),
        "perf": counts(perf_labels, perf_counter) if perf_count else {"l": ["暂无数据"], "d": [0]},
        "salhist": counts(salary_labels, salary_hist),
        "depsal": top_values(dept_salary),
        "depto": top_values(dept_turnover),
        "perfsal": top_values({k: avg(v) for k, v in perf_salary_groups.items()}),
        "move": {"l": ["入职", "离职"], "d": [join_count, leave_count]},
        "rolesal": top_values(role_salary),
        "gensal": top_values(gender_salary),
        "cityto": top_values(city_turnover),
        "roleto": top_values(role_turnover),
        "eduto": top_values(edu_turnover),
        "tentoto": top_values(tenure_turnover),
        "s5num": {"l": ["最低", "P25", "中位", "P75", "最高"], "d": [round(v, 1) for v in [salary_min, salary_p25, salary_median, salary_p75, salary_max]]},
        "edusal": top_values(edu_salary),
        "citysal": top_values(city_salary),
        "dephp": top_values({k: pct(v[0], v[1]) for k, v in dept_high_perf.items()}),
        "rolecnt": top_counter(role_counter),
        "deptcnt": top_counter(dept_counter),
        "deptA": {"l": dept_order, "d": [dept_perf_grade[d].get("A", 0) for d in dept_order]},
        "deptBp": {"l": dept_order, "d": [dept_perf_grade[d].get("B+", 0) for d in dept_order]},
        "deptB": {"l": dept_order, "d": [dept_perf_grade[d].get("B", 0) for d in dept_order]},
        "deptC": {"l": dept_order, "d": [dept_perf_grade[d].get("C", 0) for d in dept_order]},
        "saltenure": top_values(salary_by_tenure),
        "trate": {"l": ["整体离职率", "健康阈值"], "d": [round(turnover_rate, 1), 5]},
        "contractStatus": {"l": contract_labels, "d": contract_data},
        "newHireTurnover": round(new_hire_turnover_rate, 1),
    }

    gender_ratio = "暂无"
    if active_count:
        male = gender_counter.get("男", 0)
        female = gender_counter.get("女", 0)
        gender_ratio = f"{male}:{female}"

    def pct_count_rows(counter, denominator):
        return [[k, v, fmt_pct(pct(v, denominator))] for k, v in counter.most_common(10)] or [["暂无数据", 0, "0.0%"]]

    contract_detail = []
    for e in sorted(expired_contract + contract_30 + contract_90, key=lambda x: x["contract_end"] or dt.date.max)[:12]:
        days = (e["contract_end"] - report_date).days if e["contract_end"] else 0
        status = "已过期" if days < 0 else f"{days}天内到期"
        contract_detail.append([e["name"] or e["id"], e["dept"], e["contract_end"].isoformat() if e["contract_end"] else "未填", status])
    if not contract_detail:
        contract_detail = [["暂无预警", "-", "-", "正常"]]

    salary_rows_by_position = []
    for position, values in sorted(group_average(salaries, "position", "total").items(), key=lambda kv: kv[1], reverse=True)[:8]:
        vals = [s["total"] for s in salaries if s["position"] == position]
        salary_rows_by_position.append([position, fmt_money(min(vals)), fmt_money(max(vals)), fmt_money(avg(vals))])

    bl = {
        "active_count": table(["指标", "数值"], [["在职人数", f"{active_count}人"]]),
        "avg_tenure": table(["均值", "中位数"], [[f"{avg_tenure:.1f}年", f"{med_tenure:.1f}年"]]),
        "median_tenure": table(["均值", "中位数"], [[f"{avg_tenure:.1f}年", f"{med_tenure:.1f}年"]]),
        "tenure_dist": table(["司龄段", "人", "占比"], pct_count_rows(tenure_counter, active_count)),
        "probation_count": table(["试用期", "转正率"], [[f"{len(probation)}人", fmt_pct(confirm_rate)]]),
        "confirm_rate": table(["应转正", "已转正", "转正率"], [[len(confirmed_due), len(confirmed), fmt_pct(confirm_rate)]]),
        "avg_age": table(["平均年龄"], [[f"{avg_age:.1f}岁"]]),
        "age_dist": table(["年龄段", "人", "占比"], [[k, age_counter[k], fmt_pct(pct(age_counter[k], active_count))] for k in age_labels]),
        "gender_dist": table(["性别", "人", "占比"], pct_count_rows(gender_counter, active_count)),
        "edu_dist": table(["学历", "人", "占比"], pct_count_rows(edu_counter, active_count)),
        "marriage_dist": table(["婚育", "人", "占比"], pct_count_rows(marriage_counter, active_count)),
        "school_dist": table(["院校", "人"], school_counter.most_common(10) or [["暂无数据", 0]]),
        "role_dist": table(["角色", "人", "占比"], pct_count_rows(role_type_counter, active_count)),
        "position_dist": table(["岗位", "人"], role_counter.most_common(10) or [["暂无数据", 0]]),
        "dept_dist": table(["部门", "人", "占比"], pct_count_rows(dept_counter, active_count)),
        "contract_warning_rate": table(["级别", "人", "占比"], [[label, value, fmt_pct(pct(value, active_count))] for label, value in zip(contract_labels, contract_data)]),
        "contract_warnings": table(["姓名", "部门", "到期", "状态"], contract_detail),
        "join_count": table(["类型", "人"], [["入职", join_count], ["合计", join_count]]),
        "leave_count": table(["类型", "人"], [["离职", leave_count], ["合计", leave_count]]),
        "net_change": table(["入", "离", "净增"], [[join_count, leave_count, f"{net_change:+d}"]]),
        "move_type_dist": table(["类型", "人"], [["入职", join_count], ["离职", leave_count]]),
        "leave_dept_dist": table(["部门", "离职", "离职率"], [[k, Counter(e["dept"] for e in resigned)[k], fmt_pct(v)] for k, v in sorted(dept_turnover.items(), key=lambda kv: kv[1], reverse=True)[:8]] or [["暂无", 0, "0.0%"]]),
        "leave_position_dist": table(["岗位", "离职"], Counter(e["position"] for e in resigned).most_common(8) or [["暂无", 0]]),
        "turnover_formula": table(["项目", "数值"], [["本月离职", f"{leave_count}人"], ["当前在职", f"{active_count}人"], ["离职率", fmt_pct(turnover_rate)]]),
        "perf_grade_dist": table(["等级", "人", "占比"], [[g, perf_counter.get(g, 0), fmt_pct(pct(perf_counter.get(g, 0), perf_count))] for g in perf_labels]),
        "perf_coverage_rate": table(["在职", "有绩效", "覆盖率"], [[active_count, perf_count, fmt_pct(perf_coverage_rate)]]),
        "high_perf_rate": table(["A+B+", "有绩效", "占比"], [[high_perf_count, perf_count, fmt_pct(high_perf_rate)]]),
        "perf_quant_mean": table(["指标", "数值"], [["均分", f"{perf_score_mean:.1f}"]]),
        "perf_salary_cross": table(["等级", "人", "均薪"], [[g, perf_counter.get(g, 0), fmt_money(avg(perf_salary_groups.get(g, [])))] for g in perf_labels]),
        "perf_by_dept": table(["部门", "高绩效占比"], [[k, fmt_pct(pct(v[0], v[1]))] for k, v in dept_high_perf.items()] or [["暂无", "0.0%"]]),
        "perf_by_tenure": table(["司龄", "高绩效占比"], [[k, fmt_pct(pct(v[0], v[1]))] for k, v in tenure_high_perf.items()] or [["暂无", "0.0%"]]),
        "perf_by_position": table(["岗位", "高绩效占比"], [[k, fmt_pct(pct(v[0], v[1]))] for k, v in role_high_perf.items()] or [["暂无", "0.0%"]]),
        "salary_mean": table(["均值", "中位数", "P90"], [[fmt_money(salary_mean), fmt_money(salary_median), fmt_money(salary_p90)]]),
        "salary_median": table(["P50"], [[fmt_money(salary_median)]]),
        "salary_5num": table(["最低", "P25", "P50", "P75", "最高"], [[fmt_money(salary_min), fmt_money(salary_p25), fmt_money(salary_median), fmt_money(salary_p75), fmt_money(salary_max)]]),
        "salary_hist": table(["区间", "人", "占比"], [[k, salary_hist[k], fmt_pct(pct(salary_hist[k], len(salary_values)))] for k in salary_labels]),
        "salary_range": table(["最低", "最高", "带宽", "比"], [[fmt_money(salary_min), fmt_money(salary_max), fmt_money(salary_max - salary_min), f"{salary_range_ratio:.2f}"]]),
        "salary_range_ratio": table(["最低", "最高", "带宽", "比"], [[fmt_money(salary_min), fmt_money(salary_max), fmt_money(salary_max - salary_min), f"{salary_range_ratio:.2f}"]]),
        "bonus_mean": table(["奖金均值"], [[fmt_money(bonus_mean)]]),
        "bonus_total": table(["总额", "人数"], [[fmt_money(bonus_total), len(salaries)]]),
        "bonus_coverage_rate": table(["有奖金", "在职", "覆盖率"], [[len(bonus_positive), active_count, fmt_pct(bonus_coverage_rate)]]),
        "bonus_salary_ratio": table(["奖金", "薪资", "比率"], [[fmt_money(bonus_mean), fmt_money(salary_mean), fmt_pct(bonus_salary_ratio)]]),
        "salary_by_dept": table(["部门", "均值"], [[k, fmt_money(v)] for k, v in sorted(dept_salary.items(), key=lambda kv: kv[1], reverse=True)[:8]] or [["暂无", "¥0"]]),
        "salary_by_role": table(["岗位", "均值"], [[k, fmt_money(v)] for k, v in sorted(role_salary.items(), key=lambda kv: kv[1], reverse=True)[:8]] or [["暂无", "¥0"]]),
        "salary_by_edu": table(["学历", "均值"], [[k, fmt_money(v)] for k, v in sorted(edu_salary.items(), key=lambda kv: kv[1], reverse=True)[:8]] or [["暂无", "¥0"]]),
        "salary_by_city": table(["城市", "均值"], [[k, fmt_money(v)] for k, v in sorted(city_salary.items(), key=lambda kv: kv[1], reverse=True)[:8]] or [["暂无", "¥0"]]),
        "salary_by_tenure": table(["司龄", "均值"], [[k, fmt_money(v)] for k, v in sorted(salary_by_tenure.items())] or [["暂无", "¥0"]]),
        "gender_salary_gap": table(["性别", "人数", "均值"], [[k, gender_counter.get(k, 0), fmt_money(v)] for k, v in sorted(gender_salary.items(), key=lambda kv: kv[1], reverse=True)] or [["暂无", 0, "¥0"]]),
        "salary_by_position": table(["岗位", "最低", "最高", "均值"], salary_rows_by_position or [["暂无", "¥0", "¥0", "¥0"]]),
        "turnover_rate": table(["离职", "当前在职", "离职率"], [[leave_count, active_count, fmt_pct(turnover_rate)]]),
        "new_hire_turnover_rate": table(["1年内离职", "1年内人数", "离职率"], [[len(one_year_resigned), len(one_year_people), fmt_pct(new_hire_turnover_rate)]]),
        "dept_turnover": table(["部门", "离职率"], [[k, fmt_pct(v)] for k, v in sorted(dept_turnover.items(), key=lambda kv: kv[1], reverse=True)[:8]] or [["暂无", "0.0%"]]),
        "city_turnover": table(["城市", "离职率"], [[k, fmt_pct(v)] for k, v in sorted(city_turnover.items(), key=lambda kv: kv[1], reverse=True)[:8]] or [["暂无", "0.0%"]]),
        "role_turnover": table(["岗位", "离职率"], [[k, fmt_pct(v)] for k, v in sorted(role_turnover.items(), key=lambda kv: kv[1], reverse=True)[:8]] or [["暂无", "0.0%"]]),
        "edu_turnover": table(["学历", "离职率"], [[k, fmt_pct(v)] for k, v in sorted(edu_turnover.items(), key=lambda kv: kv[1], reverse=True)[:8]] or [["暂无", "0.0%"]]),
        "tenure_turnover": table(["司龄", "离职率"], [[k, fmt_pct(v)] for k, v in sorted(tenure_turnover.items(), key=lambda kv: kv[1], reverse=True)[:8]] or [["暂无", "0.0%"]]),
        "key_role_turnover_rate": table(["关键角色", "离职", "离职率"], [[len(key_roles), len(key_role_resigned), fmt_pct(key_role_turnover_rate)]]),
    }

    focus_html = "".join(
        [
            f'<div class="focus-item"><div class="focus-num">{contract_warning_count}</div><div class="focus-label">合同待续签</div><div class="focus-tag warn">待办</div></div>',
            f'<div class="focus-item"><div class="focus-num grn">{fmt_pct(new_hire_turnover_rate)}</div><div class="focus-label">新员工离职率</div><div class="focus-tag good">亮点</div></div>',
            f'<div class="focus-item"><div class="focus-num grn">{fmt_pct(high_perf_rate)}</div><div class="focus-label">高绩效占比</div><div class="focus-tag good">绩效</div></div>',
            f'<div class="focus-item"><div class="focus-num" style="color:var(--err)">{fmt_pct(turnover_rate)}</div><div class="focus-label">整体离职率</div><div class="focus-tag warn">关注</div></div>',
        ]
    )
    kpi_html = "".join(
        [
            f'<div class="kpi-c cb"><div class="kpi-v">{active_count}<span class="kpi-u">人</span></div><div class="kpi-l">在职人数</div><span class="kpi-tr fl">当前</span></div>',
            f'<div class="kpi-c co"><div class="kpi-v">{fmt_money(salary_mean)}</div><div class="kpi-l">薪资均值</div><span class="kpi-tr fl">月度</span></div>',
            f'<div class="kpi-c cr"><div class="kpi-v">{fmt_pct(high_perf_rate)}</div><div class="kpi-l">高绩效占比 (A+B+)</div><span class="kpi-tr up">{high_perf_count}人</span></div>',
            f'<div class="kpi-c ce"><div class="kpi-v">{fmt_pct(turnover_rate)}</div><div class="kpi-l">整体离职率</div><span class="kpi-tr fl">{leave_count}人</span></div>',
            f'<div class="kpi-c cp"><div class="kpi-v">{fmt_pct(contract_warning_rate)}</div><div class="kpi-l">合同预警率</div><span class="kpi-tr dn" style="background:var(--err-bg);color:var(--err)">{contract_warning_count}人</span></div>',
        ]
    )
    detail_group_html = f"""
      <div class="kg">
        <div class="kg-h"><div class="kg-dot" style="background:var(--pri)"></div>人员结构</div>
        <div class="kg-cards c4">
          <div class="kg-item"><div class="kg-item-info"><div class="kg-item-label">平均司龄</div><div class="kg-item-val">{avg_tenure:.1f}<span class="u">年</span></div><div class="kg-item-sub">中位 {med_tenure:.1f}年</div></div></div>
          <div class="kg-item"><div class="kg-item-info"><div class="kg-item-label">平均年龄</div><div class="kg-item-val">{avg_age:.1f}<span class="u">岁</span></div><div class="kg-item-sub">&nbsp;</div></div></div>
          <div class="kg-item"><div class="kg-item-info"><div class="kg-item-label">男女员工占比</div><div class="kg-item-val">{gender_ratio}</div><div class="kg-item-sub">按在职人员统计</div></div></div>
          <div class="kg-item"><div class="kg-item-info"><div class="kg-item-label">最高学历分布</div><div class="kg-item-val">{edu_counter.most_common(1)[0][0] if edu_counter else "暂无"}</div><div class="kg-item-sub">{edu_counter.most_common(1)[0][1] if edu_counter else 0}人</div></div></div>
        </div>
      </div>
      <div class="kg">
        <div class="kg-h"><div class="kg-dot" style="background:var(--grn)"></div>人员流动</div>
        <div class="kg-cards c5">
          <div class="kg-item"><div class="kg-item-info"><div class="kg-item-label">本月入职</div><div class="kg-item-val" style="color:var(--grn)">{join_count}<span class="u">人</span></div><div class="kg-item-sub">按入职日期</div></div></div>
          <div class="kg-item"><div class="kg-item-info"><div class="kg-item-label">本月离职</div><div class="kg-item-val" style="color:var(--err)">{leave_count}<span class="u">人</span></div><div class="kg-item-sub">需离职日期字段</div></div></div>
          <div class="kg-item"><div class="kg-item-info"><div class="kg-item-label">净增减</div><div class="kg-item-val" style="color:var(--grn)">{net_change:+d}<span class="u">人</span></div><div class="kg-item-sub">入-离</div></div></div>
          <div class="kg-item"><div class="kg-item-info"><div class="kg-item-label">新员工离职率</div><div class="kg-item-val" style="color:var(--grn)">{fmt_pct(new_hire_turnover_rate)}</div><div class="kg-item-sub">入职1年内</div></div></div>
          <div class="kg-item"><div class="kg-item-info"><div class="kg-item-label">关键角色离职率</div><div class="kg-item-val" style="color:var(--err)">{fmt_pct(key_role_turnover_rate)}</div><div class="kg-item-sub">关键岗位</div></div></div>
        </div>
      </div>
      <div class="kg">
        <div class="kg-h"><div class="kg-dot" style="background:var(--pnk)"></div>绩效管理</div>
        <div class="kg-cards c2">
          <div class="kg-item"><div class="kg-item-info"><div class="kg-item-label">绩效覆盖率</div><div class="kg-item-val" style="color:var(--pnk)">{fmt_pct(perf_coverage_rate)}</div><div class="kg-item-sub">{perf_count}/{active_count}人</div></div></div>
          <div class="kg-item"><div class="kg-item-info"><div class="kg-item-label">高绩效占比</div><div class="kg-item-val" style="color:var(--pnk)">{fmt_pct(high_perf_rate)}</div><div class="kg-item-sub">A+B+ · {high_perf_count}人</div></div></div>
        </div>
      </div>
      <div class="kg">
        <div class="kg-h"><div class="kg-dot" style="background:var(--gld)"></div>薪酬总览</div>
        <div class="kg-cards c2">
          <div class="kg-item"><div class="kg-item-info"><div class="kg-item-label">薪资中位数</div><div class="kg-item-val" style="color:var(--gld)">{fmt_money(salary_median)}</div><div class="kg-item-sub">P50</div></div></div>
          <div class="kg-item"><div class="kg-item-info"><div class="kg-item-label">奖金覆盖率</div><div class="kg-item-val" style="color:var(--gld)">{fmt_pct(bonus_coverage_rate)}</div><div class="kg-item-sub">有奖金</div></div></div>
        </div>
      </div>
    """

    tips_html = "".join(
        [
            f'<div class="abc yellow"><span class="abc-pri yl">关注</span><div class="abc-i">💬</div><div class="abc-t">本月离职率 {fmt_pct(turnover_rate)}</div><div class="abc-d">按当前数据计算，离职率可作为月度跟进项</div><ul class="abc-l"><li>如要准确统计，请在花名册补充离职日期</li><li>关键岗位离职需单独复盘</li></ul></div>',
            f'<div class="abc red"><span class="abc-pri cr">待办</span><div class="abc-i">📋</div><div class="abc-t">{contract_warning_count} 位伙伴合同需关注</div><div class="abc-d">含已过期、30天内到期和90天内到期</div><ul class="abc-l"><li>优先处理已过期合同</li><li>按到期日分批续签</li></ul></div>',
            f'<div class="abc blue"><span class="abc-pri bl">新人</span><div class="abc-i">🌿</div><div class="abc-t">{len(probation)} 位伙伴仍在试用期</div><div class="abc-d">转正率 {fmt_pct(confirm_rate)}</div><ul class="abc-l"><li>提前安排转正评估</li><li>关注新人适应情况</li></ul></div>',
            f'<div class="abc green"><span class="abc-pri gr">薪酬</span><div class="abc-i">💚</div><div class="abc-t">薪资均值 {fmt_money(salary_mean)}</div><div class="abc-d">奖金覆盖率 {fmt_pct(bonus_coverage_rate)}</div><ul class="abc-l"><li>持续监控部门薪资差异</li><li>重点看同岗同级公平性</li></ul></div>',
        ]
    )
    actions_html = "".join(
        [
            '<div class="actc"><span class="act-tag">人才</span><div class="act-t">把关键岗位单独列清单</div><div class="act-d">关键岗位离职和合同风险最好单独跟踪，不要只看全公司平均数。</div></div>',
            '<div class="actc"><span class="act-tag" style="background:var(--blu-l);color:var(--blu)">合同</span><div class="act-t">续签按到期日分批</div><div class="act-d">合同预警人数较多时，建议按已过期、30天内、90天内三批处理。</div></div>',
            '<div class="actc"><span class="act-tag" style="background:var(--pnk-l);color:var(--pnk)">绩效</span><div class="act-t">高绩效员工要被看见</div><div class="act-d">A/B+ 员工可作为激励、晋升和留才沟通的优先人群。</div></div>',
        ]
    )

    modules = {
        "ppl": {
            "title": "人员结构",
            "color": "#C2410C",
            "kpi": [
                {"v": f"{active_count}人", "l": "在职人数"},
                {"v": f"{avg_tenure:.1f}年", "l": "平均司龄"},
                {"v": f"{avg_age:.1f}岁", "l": "平均年龄"},
                {"v": gender_ratio, "l": "男女比"},
            ],
            "charts": [
                {"id": "mg0", "lb": "性别比例", "kind": "dn", "key": "gender"},
                {"id": "mg1", "lb": "年龄分布", "kind": "br", "key": "age", "unit": "人"},
                {"id": "mg2", "lb": "学历分布", "kind": "hb", "key": "edu", "unit": "人"},
                {"id": "mg3", "lb": "岗位结构", "kind": "hb", "key": "rolecnt", "unit": "人"},
            ],
            "data": [{"id": "tenure_dist", "lb": "司龄分布"}, {"id": "dept_dist", "lb": "部门分布"}, {"id": "role_dist", "lb": "岗位角色"}, {"id": "marriage_dist", "lb": "婚育分布"}],
            "analysis": [
                f"团队当前在职 {active_count} 人，平均司龄 {avg_tenure:.1f} 年。",
                f"平均年龄 {avg_age:.1f} 岁，男女比 {gender_ratio}。",
                f"人数最多的部门是 {dept_counter.most_common(1)[0][0] if dept_counter else '暂无'}。",
                "人员结构数据来自员工花名册，请保持员工ID唯一。",
            ],
        },
        "flow": {
            "title": "人员流动",
            "color": "#10B981",
            "kpi": [
                {"v": f"{join_count}人", "l": "本月入职"},
                {"v": f"{leave_count}人", "l": "本月离职"},
                {"v": f"{net_change:+d}人", "l": "净增减"},
                {"v": fmt_pct(turnover_rate), "l": "整体离职率"},
            ],
            "charts": [
                {"id": "mg4", "lb": "岗位离职率", "kind": "hb", "key": "roleto", "unit": "%"},
                {"id": "mg5", "lb": "城市离职率", "kind": "hb", "key": "cityto", "unit": "%"},
                {"id": "mg6", "lb": "学历离职率", "kind": "hb", "key": "eduto", "unit": "%"},
                {"id": "mg7", "lb": "司龄离职率", "kind": "hb", "key": "tentoto", "unit": "%"},
            ],
            "data": [{"id": "turnover_formula", "lb": "离职率计算"}, {"id": "leave_dept_dist", "lb": "离职部门"}, {"id": "leave_position_dist", "lb": "离职岗位"}, {"id": "new_hire_turnover_rate", "lb": "新员工离职"}],
            "analysis": [
                f"本月入职 {join_count} 人、离职 {leave_count} 人，净增 {net_change:+d} 人。",
                "当前模板未强制包含离职日期，若要月度离职率准确，请在花名册增加离职日期列。",
                f"关键角色离职率 {fmt_pct(key_role_turnover_rate)}。",
                "人员流动分析适合每月固定更新。",
            ],
        },
        "perf": {
            "title": "绩效管理",
            "color": "#EC4899",
            "kpi": [
                {"v": fmt_pct(perf_coverage_rate), "l": "绩效覆盖率"},
                {"v": fmt_pct(high_perf_rate), "l": "高绩效占比"},
                {"v": f"{perf_count}人", "l": "有绩效人数"},
                {"v": f"{perf_score_mean:.1f}", "l": "绩效均分"},
            ],
            "charts": [
                {"id": "mg8", "lb": "绩效等级占比", "kind": "sbd"},
                {"id": "mg9", "lb": "各级别人数", "kind": "br", "key": "perf", "unit": "人"},
                {"id": "mg10", "lb": "部门绩效分布", "kind": "sb"},
                {"id": "mg11", "lb": "绩效×薪资", "kind": "br", "key": "perfsal", "unit": "元"},
            ],
            "data": [{"id": "perf_grade_dist", "lb": "等级分布"}, {"id": "perf_salary_cross", "lb": "绩效×薪资"}, {"id": "perf_by_dept", "lb": "绩效×部门"}, {"id": "perf_by_position", "lb": "绩效×岗位"}],
            "analysis": [
                f"绩效覆盖率 {fmt_pct(perf_coverage_rate)}，高绩效占比 {fmt_pct(high_perf_rate)}。",
                f"绩效记录覆盖 {perf_count}/{active_count} 名在职员工。",
                "绩效数据按员工ID与花名册关联。",
                "如有多个绩效周期，系统默认取每个员工最新周期。",
            ],
        },
        "sal": {
            "title": "薪酬总览",
            "color": "#F59E0B",
            "kpi": [
                {"v": fmt_money(salary_mean), "l": "薪资均值"},
                {"v": fmt_money(salary_median), "l": "薪资中位数"},
                {"v": fmt_money(bonus_mean), "l": "奖金均值"},
                {"v": f"{salary_range_ratio:.2f}", "l": "薪酬带宽比"},
            ],
            "charts": [
                {"id": "mg12", "lb": "薪资区间分布", "kind": "br", "key": "salhist", "unit": "人"},
                {"id": "mg13", "lb": "部门薪资对比", "kind": "hb", "key": "depsal", "unit": "元"},
                {"id": "mg14", "lb": "岗位薪资对比", "kind": "hb", "key": "rolesal", "unit": "元"},
                {"id": "mg15", "lb": "性别薪资差异", "kind": "br", "key": "gensal", "unit": "元"},
            ],
            "data": [{"id": "salary_5num", "lb": "薪资五数"}, {"id": "salary_by_dept", "lb": "部门薪资"}, {"id": "salary_by_role", "lb": "岗位薪资"}, {"id": "gender_salary_gap", "lb": "性别差异"}],
            "analysis": [
                f"薪资均值 {fmt_money(salary_mean)}，中位数 {fmt_money(salary_median)}。",
                f"薪酬带宽比 {salary_range_ratio:.2f}，奖金覆盖率 {fmt_pct(bonus_coverage_rate)}。",
                "薪酬模块按员工ID与花名册关联，只统计在职员工。",
                "薪酬数据不应公开到无鉴权页面。",
            ],
        },
        "cntr": {
            "title": "合同管理",
            "color": "#3B82F6",
            "kpi": [
                {"v": fmt_pct(contract_warning_rate), "l": "合同预警率"},
                {"v": f"{len(expired_contract)}人", "l": "已过期"},
                {"v": f"{len(contract_30)}人", "l": "30天内到期"},
                {"v": f"{active_count - contract_warning_count}人", "l": "正常"},
            ],
            "charts": [{"id": "mg16", "lb": "预警状态分布", "kind": "dn", "key": "contractStatus"}],
            "data": [{"id": "contract_warning_rate", "lb": "预警率"}, {"id": "contract_warnings", "lb": "预警明细"}],
            "analysis": [
                f"合同预警人数 {contract_warning_count} 人，预警率 {fmt_pct(contract_warning_rate)}。",
                f"已过期 {len(expired_contract)} 人，30天内到期 {len(contract_30)} 人。",
                "建议按已过期、30天内、90天内分批处理。",
            ],
        },
        "tne": {
            "title": "司龄分析",
            "color": "#FB7185",
            "kpi": [
                {"v": f"{avg_tenure:.1f}年", "l": "平均司龄"},
                {"v": f"{med_tenure:.1f}年", "l": "司龄中位数"},
                {"v": fmt_pct(pct(tenure_counter.get("5-10年", 0), active_count)), "l": "5-10年占比"},
                {"v": f'{tenure_counter.get("1-3年", 0)}人', "l": "1-3年群体"},
            ],
            "charts": [
                {"id": "mg17", "lb": "司龄分布", "kind": "br", "key": "tenure", "unit": "人"},
                {"id": "mg18", "lb": "司龄离职率", "kind": "hb", "key": "tentoto", "unit": "%"},
                {"id": "mg19", "lb": "司龄薪资对比", "kind": "hb", "key": "saltenure", "unit": "元"},
            ],
            "data": [{"id": "tenure_dist", "lb": "司龄分布"}, {"id": "tenure_turnover", "lb": "司龄离职率"}, {"id": "salary_by_tenure", "lb": "司龄薪资"}, {"id": "salary_5num", "lb": "薪资五数"}],
            "analysis": [
                f"平均司龄 {avg_tenure:.1f} 年，中位数 {med_tenure:.1f} 年。",
                "司龄结构按入职日期自动计算。",
                "司龄薪资按员工ID关联薪资表计算。",
                "1-3年员工建议重点看留存和成长路径。",
            ],
        },
    }

    highlights = [
        {"num": fmt_pct(new_hire_turnover_rate), "label": "新员工离职率", "desc": "入职1年内员工留存情况"},
        {"num": f"{net_change:+d}", "label": "本月净增减", "desc": "团队规模变化"},
        {"num": fmt_pct(high_perf_rate), "label": "高绩效占比", "desc": "A+B+ 级员工占比"},
        {"num": fmt_pct(perf_coverage_rate), "label": "绩效覆盖率", "desc": "绩效数据完整度"},
        {"num": fmt_pct(pct(tenure_counter.get("5-10年", 0), active_count)), "label": "5-10年稳定员工", "desc": "核心骨干稳定性"},
        {"num": fmt_pct(bonus_coverage_rate), "label": "奖金覆盖率", "desc": "奖金激励覆盖面"},
    ]

    overview = {
        "topbarHtml": f'<strong>{report_month}</strong> <span style="color:var(--bd2)">·</span> 在职 <strong>{active_count}</strong> 人 <span style="color:var(--bd2)">·</span> 数据截止 <strong>{report_date.isoformat()}</strong>',
        "sidebarReport": f"报告: {report_month}",
        "sidebarActive": f"在职: {active_count}人",
        "greetingTitleHtml": f"早上好，今天有 <strong>{active_count}</strong> 位伙伴和你一起奋斗",
        "greetingSubtitleHtml": f"本月 <strong>{join_count}</strong> 位新朋友加入团队 · 团队净增 <strong>{net_change:+d}</strong> 人",
        "focusHtml": focus_html,
        "kpiHtml": kpi_html,
        "detailGroupHtml": detail_group_html,
        "tipsHtml": tips_html,
        "actionsHtml": actions_html,
        "footerHtml": f"<b>统计口径：</b>在职人数：统计截止日在职或离职中员工 · 司龄：入职日至截止日实际天数÷365.25 · 离职率：本月离职÷当前在职人数 · 高绩效：A+B+÷有绩效人数 · 生成时间：{dt.datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "copyText": f"HR全景看板 {report_month} 在职{active_count}人 高绩效{fmt_pct(high_perf_rate)} 离职率{fmt_pct(turnover_rate)}",
    }

    cat_notes = {
        "active_count": f"{active_count}人",
        "avg_age": f"{avg_age:.1f}岁",
        "gender_dist": gender_ratio,
        "edu_dist": f"{edu_counter.most_common(1)[0][0] if edu_counter else '暂无'}最多",
        "probation_count": f"{len(probation)}人",
        "confirm_rate": fmt_pct(confirm_rate),
        "avg_tenure": f"{avg_tenure:.1f}年",
        "median_tenure": f"{med_tenure:.1f}年",
        "contract_warning_rate": fmt_pct(contract_warning_rate),
        "join_count": f"{join_count}人",
        "leave_count": f"{leave_count}人",
        "net_change": f"{net_change:+d}人",
        "turnover_rate": fmt_pct(turnover_rate),
        "new_hire_turnover_rate": fmt_pct(new_hire_turnover_rate),
        "key_role_turnover_rate": fmt_pct(key_role_turnover_rate),
        "perf_coverage_rate": fmt_pct(perf_coverage_rate),
        "high_perf_rate": fmt_pct(high_perf_rate),
        "perf_quant_mean": f"{perf_score_mean:.1f}",
        "salary_mean": fmt_money(salary_mean),
        "salary_median": fmt_money(salary_median),
        "salary_range_ratio": f"{salary_range_ratio:.2f}",
        "bonus_mean": fmt_money(bonus_mean),
        "bonus_total": fmt_money(bonus_total),
        "bonus_coverage_rate": fmt_pct(bonus_coverage_rate),
        "bonus_salary_ratio": fmt_pct(bonus_salary_ratio),
    }

    return {"cd": cd, "bl": bl, "overview": overview, "modules": modules, "highlights": highlights, "catNotes": cat_notes}


def render_html(data: dict):
    template = TEMPLATE.read_text(encoding="utf-8-sig")
    ov = data["overview"]

    def replace_once(pattern: str, replacement: str, source: str) -> str:
        updated, count = re.subn(pattern, lambda _: replacement, source, count=1, flags=re.S)
        if count != 1:
            raise RuntimeError(f"Template section not found: {pattern[:80]}")
        return updated

    def replace_section(source: str, start_marker: str, end_marker: str, replacement: str, close_tag: str = "") -> str:
        start = source.index(start_marker)
        body_start = start + len(start_marker)
        end = source.index(end_marker, body_start)
        closing = f"\n  {close_tag}" if close_tag else ""
        return source[:body_start] + "\n    " + replacement.strip() + closing + "\n  " + source[end:]

    template = replace_once(
        r'<div class="topbar-period">.*?</div>',
        f'<div class="topbar-period">{ov["topbarHtml"]}</div>',
        template,
    )
    template = replace_once(
        r'<div class="s-item info"><span>报告: .*?</span></div>',
        f'<div class="s-item info"><span>{ov["sidebarReport"]}</span></div>',
        template,
    )
    template = replace_once(
        r'<div class="s-item info"><span>在职: .*?</span></div>',
        f'<div class="s-item info"><span>{ov["sidebarActive"]}</span></div>',
        template,
    )
    template = replace_once(r'<div class="greeting">\s*<div class="greeting-icon">.*?</div>', '<div class="greeting">\n    <div class="greeting-icon">HR</div>', template)
    template = replace_once(r'<h2>.*?</h2>', f'<h2>{ov["greetingTitleHtml"]}</h2>', template)
    template = replace_once(r'<p>.*?</p>', f'<p>{ov["greetingSubtitleHtml"]}</p>', template)
    template = replace_section(template, '  <div class="focus-row">', '\n\n  <div class="sh"><div class="st">核心指标</div></div>', ov["focusHtml"], "</div>")
    template = replace_section(template, '  <div class="kpi-g">', '\n\n  <div class="expand-bar"', ov["kpiHtml"], "</div>")
    template = replace_section(template, '  <div class="kpi-sg">', '\n  </div>\n\n  <div class="chart-sec">', ov["detailGroupHtml"])
    template = replace_section(template, '  <div class="ab-g">', '\n\n  <div class="sh"><div class="st">给HR的小建议</div></div>', ov["tipsHtml"], "</div>")
    template = replace_section(template, '  <div class="act-g">', '\n\n  <div class="fn">', ov["actionsHtml"], "</div>")
    template = replace_once(r'<div class="fn">.*?</div>', f'<div class="fn">{ov["footerHtml"]}</div>', template)

    template = replace_once(r'var CD=\{.*?\};\n\nvar CAT=', f'var CD={json.dumps(data["cd"], ensure_ascii=False)};\n\nvar CAT=', template)
    template = replace_once(r'var BL=\{.*?\n\};\n\nvar sel=', f'var BL={json.dumps(data["bl"], ensure_ascii=False)};\n\nvar sel=', template)
    for metric_id, note in data.get("catNotes", {}).items():
        pattern = r'(\{id:"' + re.escape(metric_id) + r'",label:"[^"]+",note:)"[^"]*"'
        template = re.sub(pattern, lambda m, note=note: m.group(1) + json.dumps(note, ensure_ascii=False), template)

    module_fn = """
function renderModulePage(tab){
  var p=(window.HR_DASHBOARD_DATA.modules || {})[tab]; if(!p)return;
  var kpiHtml=p.kpi.map(function(k){return'<div class="mp-kpi-c"><div class="v">'+k.v+'</div><div class="l">'+k.l+'</div></div>'}).join('');
  document.getElementById('mpKpi').innerHTML='<h3><div class="md" style="background:'+p.color+';box-shadow:0 0 6px '+p.color+'"></div>'+p.title+' — 核心指标</h3><div class="mp-kpi-g">'+kpiHtml+'</div>';
  document.getElementById('mpChart').innerHTML='<h3>趋势图表</h3><div class="mp-chart-g" id="mpChartGrid"></div>';
  var cg=document.getElementById('mpChartGrid'); cg.innerHTML='';
  p.charts.forEach(function(c,i){
    var w=document.createElement('div');
    w.innerHTML='<div class="ci-l">'+c.lb+'</div><div class="cw"><canvas id="'+c.id+'"></canvas></div>';
    cg.appendChild(w);
    setTimeout(function(){
      var s=c.key ? (CD[c.key] || {l:['暂无数据'],d:[0]}) : null;
      if(c.kind==='dn') dn(c.id,s.l,s.d);
      else if(c.kind==='br') br(c.id,s.l,s.d,c.unit||'');
      else if(c.kind==='hb') hb(c.id,s.l,s.d,c.unit||'');
      else if(c.kind==='sb') sb(c.id);
      else if(c.kind==='sbd') sbd(c.id);
    },100+i*50);
  });
  var dataHtml=p.data.map(function(d){
    var b=BL[d.id]||'<div style="color:var(--text4);font-size:12px">暂无数据</div>';
    return '<div class="dc"><div class="dc-t">'+d.lb+'</div>'+b+'</div>';
  }).join('');
  document.getElementById('mpData').innerHTML='<h3>明细数据</h3><div class="mp-data-g">'+dataHtml+'</div>';
  var anaHtml=p.analysis.map(function(a){return'<li>'+a+'</li>'}).join('');
  document.getElementById('mpAnalysis').innerHTML='<h3>详细解读</h3><div class="mp-analysis"><ul>'+anaHtml+'</ul></div>';
}
    """
    template = replace_once(r'function renderModulePage\(tab\)\{.*?\n\}\n\nfunction toggleDetail', module_fn + "\nfunction toggleDetail", template)

    block = f"""
/* HR_DASHBOARD_AUTOGEN_START */
(function(){{
  window.HR_DASHBOARD_DATA = {json.dumps(data, ensure_ascii=False)};
  Object.assign(CD, window.HR_DASHBOARD_DATA.cd || {{}});
  Object.assign(BL, window.HR_DASHBOARD_DATA.bl || {{}});
  var notes = window.HR_DASHBOARD_DATA.catNotes || {{}};
  if (window.CAT) CAT.forEach(function(group){{group.metrics.forEach(function(metric){{if(notes[metric.id]) metric.note = notes[metric.id];}})}});

  window.applyDashboardData = function(){{
    var data = window.HR_DASHBOARD_DATA;
    var ov = data.overview || {{}};
    var el;
    el = document.querySelector('.topbar-period'); if (el) el.innerHTML = ov.topbarHtml || el.innerHTML;
    var info = document.querySelectorAll('.sidebar-end .s-item.info span');
    if (info[0]) info[0].textContent = ov.sidebarReport || info[0].textContent;
    if (info[1]) info[1].textContent = ov.sidebarActive || info[1].textContent;
    el = document.querySelector('.greeting h2'); if (el) el.innerHTML = ov.greetingTitleHtml || el.innerHTML;
    el = document.querySelector('.greeting p'); if (el) el.innerHTML = ov.greetingSubtitleHtml || el.innerHTML;
    el = document.querySelector('.focus-row'); if (el && ov.focusHtml) el.innerHTML = ov.focusHtml;
    el = document.querySelector('.kpi-g'); if (el && ov.kpiHtml) el.innerHTML = ov.kpiHtml;
    el = document.querySelector('.kpi-sg'); if (el && ov.detailGroupHtml) el.innerHTML = ov.detailGroupHtml;
    el = document.querySelector('.ab-g'); if (el && ov.tipsHtml) el.innerHTML = ov.tipsHtml;
    el = document.querySelector('.act-g'); if (el && ov.actionsHtml) el.innerHTML = ov.actionsHtml;
    el = document.querySelector('.fn'); if (el && ov.footerHtml) el.innerHTML = ov.footerHtml;
  }};

  window.copyData = function(){{
    var text = (window.HR_DASHBOARD_DATA.overview || {{}}).copyText || 'HR全景看板';
    if(navigator.clipboard){{navigator.clipboard.writeText(text).then(function(){{alert('已复制')}})}}
  }};

  window.sbd = function(cn){{
    var ctx=document.getElementById(cn); if(!ctx)return;
    var total=(CD.perf.d||[]).reduce(function(a,b){{return a+b}},0);
    if(!total){{ dn(cn, ['暂无数据'], [0]); return; }}
    var g=['#EC4899','#F472B6','#F9A8D4','#FBCFE8'];
    var labels=CD.perf.l;
    var pcts=CD.perf.d.map(function(v){{return (v/total*100).toFixed(1)+'%'}});
    var ci=new Chart(ctx,{{type:'doughnut',data:{{labels:labels.map(function(l,i){{return l+' '+pcts[i]+' ('+CD.perf.d[i]+'人)'}}),datasets:[{{data:CD.perf.d,backgroundColor:g,borderWidth:2,borderColor:'#FFF7ED'}}]}},options:{{responsive:true,maintainAspectRatio:false,cutout:'55%',plugins:{{legend:{{position:'right',labels:{{font:{{size:10}},color:'#1C1917',boxWidth:10,padding:4}}}},datalabels:{{display:true,font:{{size:10,weight:'bold'}},color:'#fff',formatter:function(v){{return v}}}}}}}}}});
    chartInst.push(ci);
  }};

  window.renderModulePage = function(tab){{
    var p=(window.HR_DASHBOARD_DATA.modules || {{}})[tab]; if(!p)return;
    var kpiHtml=p.kpi.map(function(k){{return'<div class="mp-kpi-c"><div class="v">'+k.v+'</div><div class="l">'+k.l+'</div></div>'}}).join('');
    document.getElementById('mpKpi').innerHTML='<h3><div class="md" style="background:'+p.color+';box-shadow:0 0 6px '+p.color+'"></div>'+p.title+' — 核心指标</h3><div class="mp-kpi-g">'+kpiHtml+'</div>';
    document.getElementById('mpChart').innerHTML='<h3>趋势图表</h3><div class="mp-chart-g" id="mpChartGrid"></div>';
    var cg=document.getElementById('mpChartGrid'); cg.innerHTML='';
    p.charts.forEach(function(c,i){{
      var w=document.createElement('div');
      w.innerHTML='<div class="ci-l">'+c.lb+'</div><div class="cw"><canvas id="'+c.id+'"></canvas></div>';
      cg.appendChild(w);
      setTimeout(function(){{
        var s=c.key ? (CD[c.key] || {{l:['暂无数据'],d:[0]}}) : null;
        if(c.kind==='dn') dn(c.id,s.l,s.d);
        else if(c.kind==='br') br(c.id,s.l,s.d,c.unit||'');
        else if(c.kind==='hb') hb(c.id,s.l,s.d,c.unit||'');
        else if(c.kind==='sb') sb(c.id);
        else if(c.kind==='sbd') sbd(c.id);
      }},100+i*50);
    }});
    var dataHtml=p.data.map(function(d){{
      var b=BL[d.id]||'<div style="color:var(--text4);font-size:12px">暂无数据</div>';
      return '<div class="dc"><div class="dc-t">'+d.lb+'</div>'+b+'</div>';
    }}).join('');
    document.getElementById('mpData').innerHTML='<h3>明细数据</h3><div class="mp-data-g">'+dataHtml+'</div>';
    var anaHtml=p.analysis.map(function(a){{return'<li>'+a+'</li>'}}).join('');
    document.getElementById('mpAnalysis').innerHTML='<h3>详细解读</h3><div class="mp-analysis"><ul>'+anaHtml+'</ul></div>';
  }};

  window.generateHighlights = function(){{
    var grid=document.getElementById('hlGrid'); if(!grid)return;
    var items=(window.HR_DASHBOARD_DATA.highlights || []).slice(0,6);
    grid.innerHTML=items.map(function(it){{
      return '<div class="hl-item"><div class="hl-num">'+it.num+'</div><div class="hl-label">'+it.label+'</div><div class="hl-desc">'+it.desc+'</div></div>';
    }}).join('');
    var summary=document.getElementById('hlSummary');
    if(summary) summary.innerHTML='<strong>HR价值提炼：</strong>本月人员数据已按最新Excel自动刷新，可用于例会快速查看。';
  }};
}})();
/* HR_DASHBOARD_AUTOGEN_END */
"""
    needle = "renderDrawerContent();buildCharts();buildAcc();generateHighlights();"
    if needle not in template:
        raise RuntimeError("Template initialization call not found")
    html_out = template.replace(needle, block + "\napplyDashboardData();" + needle)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(html_out, encoding="utf-8")


def main():
    data = build_dashboard()
    render_html(data)
    print(f"Generated {OUTPUT}")
    print(f"Active employees: {len(load_rows(ROSTER_FILE))}")


if __name__ == "__main__":
    main()
